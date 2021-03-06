#########################################################################
#    _________        ___.                   ______________________  ___
#    \_   ___ \___.__.\_ |__   ___________  /  _____/\______   \   \/  /
#    /    \  \<   |  | | __ \_/ __ \_  __ \/   \  ___ |       _/\     /
#    \     \___\___  | | \_\ \  ___/|  | \/\    \_\  \|    |   \/     \
#     \______  / ____| |___  /\___  >__|    \______  /|____|_  /___/\  \
#            \/\/          \/     \/               \/        \/      \_/
#
#

import os
import json
import requests
import click
from openpyxl import Workbook
from tqdm import tqdm
from utils import sheet_writer
from glom import glom, Coalesce


def tag_categorization(tagging_prefix):
    return lambda value: ", ".join(
        [v.replace(tagging_prefix, "", 1).strip() for v in value if v.startswith(tagging_prefix)]
    )


THIRD_PARTY_TABLE = "Third Parties"
GAPS_TABLE = "Control Gaps (Findings)"
CONTROL_SCORES = "Control Scores"
COMPANY_TAGS = "Company Tags"

TP_COLUMNS = [
    ["Company Name", "name", "blue"],
    ["Company URL", "primary_url", "blue"],
    ["Business Unit", "business_unit", "red"],
    ["Vendor Owner", "vendor_owner", "red"],
    ["Regulation", "regulation", "red"],
    ["Likelihood", "likelihood_label", "orange"],
    ["Likelihood Value", "likelihood_score", "orange"],
    ["Impact", "impact_label", "orange"],
    ["Impact Value", "impact_score", "orange"],
    ["Assessment State", "assessment_status"],
    ["Assessment Progress", "assessment_progress"],
    ["Report order status", "subscription_status"],
    ["Report tier", "subscription_tier"],
    ["Report available", "subscription_available"],
    ["Industry", "industry"],
]

TP_MAPPING = {
    "likelihood_label": Coalesce("inherent_risk.likelihood_label", default=None),
    "likelihood_score": Coalesce("inherent_risk.likelihood_score", default=None),
    "impact_label": Coalesce("inherent_risk.impact_label", default=None),
    "impact_score": Coalesce("inherent_risk.impact_score", default=None),
    "assessment_status": Coalesce("assessment.status", default=None),
    "assessment_progress": Coalesce("assessment.progress", default=None),
    "subscription_status": Coalesce("subscription.status", default=None),
    "subscription_tier": Coalesce("subscription.tier", default=None),
    "subscription_available": Coalesce("subscription.is_report_available", default=None),
    "business_unit": (Coalesce("tags", default=[]), tag_categorization("BU:")),
    "vendor_owner": (Coalesce("tags", default=[]), tag_categorization("VO:")),
    "regulation": (Coalesce("tags", default=[]), tag_categorization("REG:")),
}

GAPS_COLUMNS = [
    ["Company Name", "company_name", "blue"],
    ["Control Name", "name", "orange"],
    ["Control Number", "number", "orange"],
    ["Level", "impact_level", "orange"],
    ["Remedy", "remedy", "orange"],
]

SCORE_COLUMNS = [
    ["Company Name", "company_name", "blue"],
    ["Control Name", "name", "blue"],
    ["Control Number", "number", "blue"],
    ["Answer State", "answer_state", "orange"],
    ["Effectiveness Score", "effectiveness_score", "orange"],
    ["Coverage Score", "coverage_score", "orange"],
    ["Maturity Score", "maturity_score", "orange"],
]

SCORE_MAPPING = {
    "effectiveness_score": Coalesce("effectiveness_score", default=None),
    "coverage_score": Coalesce("coverage_score", default=None),
    "maturity_score": Coalesce("maturity_score", default=None),
    "answer_state": Coalesce("answer_state", default=None),
}

TAG_COLUMNS = [
    ["Company Name", "company_name", "blue"],
    ["Tag", "tag"],
]


@click.command()
@click.argument("filename", required=False, default="ecosystem.xlsx")
def export_ecosystem(filename):
    api = os.environ.get("CYBERGRX_API", "https://api.cybergrx.com").rstrip("/")
    token = os.environ.get("CYBERGRX_API_TOKEN", None)
    if not token:
        raise Exception("The environment variable CYBERGRX_API_TOKEN must be set")

    uri = api + "/bulk-v1/third-parties"
    print("Fetching third parties from " + uri + " this can take some time.")
    response = requests.get(uri, headers={"Authorization": token.strip()})
    result = json.loads(response.content.decode("utf-8"))

    print("Retrieved " + str(len(result)) + " third parties from your ecosystem, building an excel.")

    wb = Workbook()
    wb["Sheet"].title = THIRD_PARTY_TABLE
    wb.create_sheet(GAPS_TABLE)
    wb.create_sheet(CONTROL_SCORES)
    wb.create_sheet(COMPANY_TAGS)

    third_party_writer = sheet_writer(wb, THIRD_PARTY_TABLE, TP_COLUMNS, mapping=TP_MAPPING)
    findings_writer = sheet_writer(wb, GAPS_TABLE, GAPS_COLUMNS)
    scores_writer = sheet_writer(wb, CONTROL_SCORES, SCORE_COLUMNS, mapping=SCORE_MAPPING)
    tags_writer = sheet_writer(wb, COMPANY_TAGS, TAG_COLUMNS)

    for tp in tqdm(result, total=len(result), desc="Third Party"):
        third_party_writer(tp)
        for tag in glom(tp, Coalesce("tags", default=[])):
            tags_writer({"tag": tag, "company_name": tp["name"]})

        for finding in glom(tp, Coalesce("residual_risk.findings", default=[])):
            finding["company_name"] = tp["name"]
            findings_writer(finding)

        for score in glom(tp, Coalesce("residual_risk.scores", default=[])):
            score["company_name"] = tp["name"]
            scores_writer(score)

    # Finalize each writer (fix width, ETC)
    third_party_writer.finalizer()
    findings_writer.finalizer()
    scores_writer.finalizer()
    tags_writer.finalizer()
    wb.save("ecosystem.xlsx")


if __name__ == "__main__":
    export_ecosystem()
