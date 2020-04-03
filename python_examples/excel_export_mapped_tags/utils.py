#########################################################################
#    _________        ___.                   ______________________  ___
#    \_   ___ \___.__.\_ |__   ___________  /  _____/\______   \   \/  /
#    /    \  \<   |  | | __ \_/ __ \_  __ \/   \  ___ |       _/\     /
#    \     \___\___  | | \_\ \  ___/|  | \/\    \_\  \|    |   \/     \
#     \______  / ____| |___  /\___  >__|    \______  /|____|_  /___/\  \
#            \/\/          \/     \/               \/        \/      \_/
#
#

import os
import json
import requests
from openpyxl import Workbook
from openpyxl.styles.fills import FILL_SOLID
from openpyxl.styles import Color, PatternFill, Font, Border, Side
from openpyxl.styles import colors
from openpyxl.cell import Cell
from tqdm import tqdm
from glom import glom


def _cell_value(cell):
    return "{}".format(cell.value).strip() if cell and cell.value else ""


def sheet_writer(wb, name, columns, mapping=None):
    if not mapping:
        mapping = {}

    for c in columns:
        if not mapping.get(c[1], None):
            mapping[c[1]] = c[1]

    def builder(sheet):
        for idx, injector in enumerate(columns):
            cell = sheet.cell(row=1, column=1 + idx)
            cell.value = injector[0]
            cell.font = cell.font.copy(bold=True)

            if len(injector) <= 2:
                cell.fill = PatternFill(FILL_SOLID, start_color="C9C9C9", end_color="C9C9C9")
            elif injector[2] == "red":
                cell.fill = PatternFill(FILL_SOLID, start_color="B35651", end_color="B35651")
                cell.font = cell.font.copy(color=colors.WHITE)
            elif injector[2] == "blue":
                cell.fill = PatternFill(FILL_SOLID, start_color="0065B8", end_color="0065B8")
                cell.font = cell.font.copy(color=colors.WHITE)
            elif injector[2] == "orange":
                cell.fill = PatternFill(FILL_SOLID, start_color="FFB802", end_color="FFB802")
            else:
                cell.fill = PatternFill(FILL_SOLID, start_color="C9C9C9", end_color="C9C9C9")

        def write_value(_row, _col, _val):
            cell = sheet.cell(row=_row, column=_col)
            cell.value = _val

        __non_local = {"row": 2}

        def writer(blob):
            transformed = glom(blob, mapping)
            multi_row = 0
            for idx, injector in enumerate(columns):
                value = transformed[injector[1]]
                if value is None:
                    continue

                if not isinstance(value, (list, tuple)):
                    write_value(__non_local["row"], 1 + idx, value)
                else:
                    multi_row = max(multi_row, len(value))
                    for i, v in enumerate(value):
                        write_value(__non_local["row"] + i, 1 + idx, v)

            __non_local["row"] = __non_local["row"] + (multi_row if multi_row else 1)

        def finalizer():
            for column_cells in sheet.columns:
                length = min(125, max(9, max(len(_cell_value(cell)) + 1 for cell in column_cells)))

                for cell in column_cells:
                    cell.alignment = cell.alignment.copy(wrapText=True)

                sheet.column_dimensions[column_cells[0].column_letter].width = length

        writer.finalizer = finalizer
        return writer

    return builder(wb[name])
