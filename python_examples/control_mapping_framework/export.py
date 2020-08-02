#########################################################################
#    _________        ___.                   ______________________  ___
#    \_   ___ \___.__.\_ |__   ___________  /  _____/\______   \   \/  /
#    /    \  \<   |  | | __ \_/ __ \_  __ \/   \  ___ |       _/\     /
#    \     \___\___  | | \_\ \  ___/|  | \/\    \_\  \|    |   \/     \
#     \______  / ____| |___  /\___  >__|    \______  /|____|_  /___/\  \
#            \/\/          \/     \/               \/        \/      \_/
#
#

import json
import os
import re
import shutil
from urllib.parse import quote

import click
import requests
import xlwings as xw
from config import (
    YESTERDAY,
    CONTROL_SCORES,
    GAPS_TABLE,
    COMPANY_TAGS,
    TP_COLUMNS,
    TP_MAPPING,
    GAPS_COLUMNS,
    SCORE_COLUMNS,
    SCORE_MAPPING,
    TAG_COLUMNS,
    GAPS_SUMMARY,
    THIRD_PARTY_TABLE,
)
from ecosystem_utils import init_ecosystem_writer
from excel_utils import process_excel_template
from glom import glom, Coalesce
from openpyxl import load_workbook
from openpyxl.cell import MergedCell
from reporting import create_report
from tqdm import tqdm
from utils import sheet_writer, control_search, create_sheet


def init_workbook(filename):
    wb = load_workbook(filename=filename)

    main = next((s for _, s in enumerate(wb)))
    main.title = "Mapped Controls"
    insert_controls = set()
    for row in main:
        insert_controls.update(control_search({idx: col for idx, col in enumerate(row)}))

    create_sheet(wb, CONTROL_SCORES)
    create_sheet(wb, GAPS_TABLE)
    create_sheet(wb, COMPANY_TAGS)
    create_sheet(wb, THIRD_PARTY_TABLE)

    findings_writer = sheet_writer(wb, GAPS_TABLE, GAPS_COLUMNS)
    scores_writer = sheet_writer(
        wb, CONTROL_SCORES, SCORE_COLUMNS, mapping=SCORE_MAPPING, insert_controls=insert_controls
    )
    tags_writer = sheet_writer(wb, COMPANY_TAGS, TAG_COLUMNS)
    third_party_writer = sheet_writer(wb, THIRD_PARTY_TABLE, TP_COLUMNS, mapping=TP_MAPPING)

    return wb, scores_writer, findings_writer, tags_writer, third_party_writer


def finalize_workbook(wb, excel_filename, debug=False):
    temporary_filename = "temporary-workbook.xlsx"
    if os.path.exists(temporary_filename):
        os.remove(temporary_filename)

    # Save the raw file as a temporary file so all of the formulas can be calculated
    wb.save(temporary_filename)

    try:
        # Open the workbook so Excel can compute all formulas and store them in the document
        temp_wb = xw.Book(temporary_filename)
        temp_wb.save()
        temp_wb.close()

        final_workbook = load_workbook(filename=temporary_filename, data_only=True)

        # In not in debug mode remove computed values
        if not debug:
            # For every cell in the document write the value, this will be the computed formula because we opened using data_only=True
            for _, sheet in enumerate(final_workbook):
                for row in sheet:
                    for cell in row:
                        if isinstance(cell, MergedCell):
                            continue
                        cell.value = cell.value

            # Remove supporting sheets
            del final_workbook[CONTROL_SCORES]
            del final_workbook[GAPS_TABLE]
            del final_workbook[COMPANY_TAGS]

        # Save a final copy
        if os.path.exists(excel_filename):
            os.remove(excel_filename)
        final_workbook.save(excel_filename)
    finally:
        # Clean up after ourselves
        if os.path.exists(temporary_filename):
            os.remove(temporary_filename)


@click.command()
@click.option(
    "--excel-template-name",
    help="Filename of the controls mapping template",
    required=False,
    default="excel-template.xlsx",
)
@click.option(
    "--report-template-name", help="Filename of the report template", required=False, default="report-template.docx",
)
@click.option(
    "--reports-from",
    help="Retrieve reports that are 'newer' than this date, defaults to yesterday",
    required=False,
    default=YESTERDAY,
)
@click.option(
    "--ecosystem-template", help="Produce an ecosystem level excel report from this template", required=False,
)
@click.option(
    "--excel-report",
    help="Treat the excel template as a Jinja template instead of producing a Word document",
    is_flag=True,
)
@click.option(
    "--debug", help="Put the script into debug mode, extra data will be preserved in this mode", is_flag=True,
)
def map_analytics(excel_template_name, report_template_name, reports_from, ecosystem_template, excel_report, debug):
    if not os.path.exists(excel_template_name):
        raise Exception(f"The --excel-template-name={excel_template_name} does not exist")

    if not excel_report and not os.path.exists(report_template_name):
        raise Exception(f"The --report-template-name={report_template_name} does not exist")

    for f in [f for f in os.listdir(".") if os.path.isfile(f)]:
        if f in [excel_template_name, report_template_name, ecosystem_template]:
            continue

        if os.path.splitext(f)[1] in [".xlsx", ".docx", ".json"]:
            print(f"Cleaning up old report {f}")
            os.remove(f)

    api = os.environ.get("CYBERGRX_API", "https://api.cybergrx.com").rstrip("/")
    token = os.environ.get("CYBERGRX_API_TOKEN", None)
    if not token:
        raise Exception("The environment variable CYBERGRX_API_TOKEN must be set")

    ecosystem_writer = init_ecosystem_writer(ecosystem_template)

    uri = f"{api}/bulk-v1/third-parties?report_date={quote(reports_from)}"
    print(f"Fetching third parties from {uri} this can take some time.")
    response = requests.get(uri, headers={"Authorization": token.strip()})
    result = json.loads(response.content.decode("utf-8"))

    print(f"Retrieved {str(len(result))} third parties from your ecosystem, building an excel.")
    for tp in tqdm(result, total=len(result), desc="Third Party"):
        company_name = tp["name"]
        report_date = glom(tp, Coalesce("residual_risk.date", default=""))

        scores = glom(tp, Coalesce("residual_risk.scores", default=[]))
        if not scores:
            continue

        tier = glom(tp, Coalesce("residual_risk.tier", default=0))
        if tier not in [1, 2]:
            print(f"{company_name} had a T{tier} report, this tier is not supported.")
            continue

        # Inject gaps summary into the TP
        tp.update(glom(tp, Coalesce(GAPS_SUMMARY, default={})))

        wb, scores_writer, findings_writer, tags_writer, third_party_writer = init_workbook(excel_template_name)

        for tag in glom(tp, Coalesce("tags", default=[])):
            tag_meta = {"tag": tag, "company_name": company_name}
            tags_writer(tag_meta)
            ecosystem_writer.tags_writer(tag_meta)

        for finding in glom(tp, Coalesce("residual_risk.findings", default=[])):
            finding["company_name"] = company_name
            findings_writer(finding)
            ecosystem_writer.findings_writer(finding)

        for score in scores:
            score["company_name"] = company_name
            scores_writer(score)
            ecosystem_writer.scores_writer(score)

        # Write third party metadata
        third_party_writer(tp)
        ecosystem_writer.third_party_writer(tp)

        # Finalize each writer (fix width, ETC)
        findings_writer.finalizer()
        scores_writer.finalizer()
        tags_writer.finalizer()
        third_party_writer.finalizer()

        output_filename = f'{re.sub("[^A-Za-z0-9 &]+", "", company_name).replace(" ", "-")}_{report_date}'
        excel_filename = f"{output_filename}.xlsx"

        finalize_workbook(wb, excel_filename, debug=debug)
        if excel_report:
            process_excel_template(excel_filename, metadata=tp, debug=debug)
        else:
            create_report(excel_filename, report_template_name, f"{output_filename}.docx", metadata=tp, debug=debug)

        ecosystem_writer.procecss_excel(excel_filename, company_name)

    ecosystem_writer.finalizer()


@click.command()
@click.option(
    "--excel-template-name",
    help="Filename of the controls mapping template",
    required=False,
    default="excel-template.xlsx",
)
@click.option(
    "--debug-json", help="Process this debug JSON and create a new report", required=True,
)
def run_excel_template(excel_template_name, debug_json):
    if not os.path.exists(excel_template_name):
        raise Exception(f"{excel_template_name} does not exist.")

    if not os.path.exists(debug_json):
        raise Exception(f"{debug_json} does not exist.")

    with open(debug_json) as f:
        metadata = json.load(f)

    json_filename = os.path.basename(debug_json)
    excel_filename = f"{os.path.splitext(json_filename)[0]}.xlsx"

    shutil.copyfile(excel_template_name, excel_filename)

    process_excel_template(excel_filename, metadata=metadata, debug=True)


@click.command()
@click.option(
    "--report-template-name", help="Filename of the report template", required=False, default="report-template.docx",
)
@click.option(
    "--excel-report-name", help="Process this excel report and generate a word document", required=True,
)
def excel_to_report(excel_report_name, report_template_name):
    file_name = os.path.basename(excel_report_name)

    metadata = {}
    json_file = f"{os.path.splitext(file_name)[0]}.json"
    if os.path.exists(json_file):
        with open(json_file) as f:
            metadata = json.load(f)

    create_report(
        excel_report_name,
        report_template_name,
        f"{os.path.splitext(file_name)[0]}.docx",
        metadata=metadata,
        debug=False,
    )


@click.command()
@click.option(
    "--excel-template-name",
    help="Filename of the controls mapping template",
    required=False,
    default="excel-template.xlsx",
)
def test_excel_template(excel_template_name):
    if not os.path.exists(excel_template_name):
        raise Exception(f"The --excel-template-name={excel_template_name} does not exist")

    init_workbook(excel_template_name)


@click.group()
def cli():
    pass


cli.add_command(map_analytics)
cli.add_command(run_excel_template)
cli.add_command(excel_to_report)
cli.add_command(test_excel_template)


if __name__ == "__main__":
    cli()
