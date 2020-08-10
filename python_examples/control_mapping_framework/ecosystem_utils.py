#########################################################################
#    _________        ___.                   ______________________  ___
#    \_   ___ \___.__.\_ |__   ___________  /  _____/\______   \   \/  /
#    /    \  \<   |  | | __ \_/ __ \_  __ \/   \  ___ |       _/\     /
#    \     \___\___  | | \_\ \  ___/|  | \/\    \_\  \|    |   \/     \
#     \______  / ____| |___  /\___  >__|    \______  /|____|_  /___/\  \
#            \/\/          \/     \/               \/        \/      \_/
#
#
from attrdict import AttrDict
from config import (
    CONTROL_SCORES,
    GAPS_TABLE,
    COMPANY_TAGS,
    TP_COLUMNS,
    TP_MAPPING,
    GAPS_COLUMNS,
    SCORE_COLUMNS,
    SCORE_MAPPING,
    TAG_COLUMNS,
    MAPPED_CONTROLS_TABLE,
    THIRD_PARTY_TABLE,
    RESIDUAL_RISK_TABLE,
    RESIDUAL_RISK_COLUMNS,
)
from glom import glom, Coalesce
from openpyxl import load_workbook
from stringcase import snakecase
from utils import sheet_writer, create_sheet, cell_value


def read_ecosystem_template(wb):
    try:
        sheet = wb[MAPPED_CONTROLS_TABLE]
    except KeyError:
        sheet = next((s for _, s in enumerate(wb)))
        sheet.title = MAPPED_CONTROLS_TABLE

    mapping = {}
    for idx, header in enumerate(next(sheet.iter_rows())):
        mapping[snakecase(cell_value(header).lower())] = idx + 1

    if "company_name" not in mapping:
        raise Exception(f"There is no Company Name column in {MAPPED_CONTROLS_TABLE}")

    row_idx = 2

    def process_excel(excel_filename, company_name):
        nonlocal row_idx
        source_wb = load_workbook(excel_filename)
        source_sheet = source_wb[MAPPED_CONTROLS_TABLE]

        rowiterator = iter(source_sheet.rows)
        source_mapping = {}
        for _idx, _header in enumerate(next(rowiterator)):
            source_mapping[_idx] = snakecase(cell_value(_header).lower())

        for row in rowiterator:
            sheet.cell(row=row_idx, column=mapping["company_name"]).value = company_name
            for i, col in enumerate(row):
                sheet.cell(row=row_idx, column=mapping[source_mapping[i]]).value = col.value
            row_idx += 1

    return process_excel


def init_ecosystem_writer(ecosystem_template):
    if not ecosystem_template:
        return AttrDict(
            {
                "tags_writer": lambda tag_meta: False,
                "findings_writer": lambda finding: False,
                "scores_writer": lambda score: False,
                "third_party_writer": lambda tp: False,
                "process_excel": lambda excel_filename, company_name: False,
                "finalizer": lambda: False,
            }
        )

    wb = load_workbook(ecosystem_template)
    process_excel = read_ecosystem_template(wb)
    create_sheet(wb, CONTROL_SCORES)
    create_sheet(wb, GAPS_TABLE)
    create_sheet(wb, COMPANY_TAGS)
    create_sheet(wb, THIRD_PARTY_TABLE)
    create_sheet(wb, RESIDUAL_RISK_TABLE)

    score_mapping = {"company_name": "company_name"}
    score_mapping.update(SCORE_MAPPING)
    score_columns = [["Company Name", "company_name"]]
    score_columns.extend(SCORE_COLUMNS)
    scores_writer = sheet_writer(wb, CONTROL_SCORES, score_columns, mapping=score_mapping)
    findings_writer = sheet_writer(wb, GAPS_TABLE, GAPS_COLUMNS)
    tags_writer = sheet_writer(wb, COMPANY_TAGS, TAG_COLUMNS)
    third_party_writer = sheet_writer(wb, THIRD_PARTY_TABLE, TP_COLUMNS, mapping=TP_MAPPING)
    residual_risk_writer = sheet_writer(wb, RESIDUAL_RISK_TABLE, RESIDUAL_RISK_COLUMNS)

    def process_third_party(tp):
        third_party_writer(tp)
        for outcome in glom(tp, Coalesce("residual_risk.residual_risk_outcomes", default=[])):
            outcome["company_name"] = tp["name"]
            residual_risk_writer(outcome)

    def finalizer():
        # Finalize each writer (fix width, ETC)
        findings_writer.finalizer()
        scores_writer.finalizer()
        tags_writer.finalizer()
        third_party_writer.finalizer()
        residual_risk_writer.finalizer()
        wb.save(filename="ecosystem.xlsx")

    return AttrDict(
        {
            "tags_writer": lambda tag_meta: tags_writer(tag_meta),
            "findings_writer": lambda finding: findings_writer(finding),
            "scores_writer": lambda score: scores_writer(score),
            "third_party_writer": process_third_party,
            "process_excel": process_excel,
            "finalizer": finalizer,
        }
    )
