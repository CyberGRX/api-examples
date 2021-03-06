#########################################################################
#    _________        ___.                   ______________________  ___
#    \_   ___ \___.__.\_ |__   ___________  /  _____/\______   \   \/  /
#    /    \  \<   |  | | __ \_/ __ \_  __ \/   \  ___ |       _/\     /
#    \     \___\___  | | \_\ \  ___/|  | \/\    \_\  \|    |   \/     \
#     \______  / ____| |___  /\___  >__|    \______  /|____|_  /___/\  \
#            \/\/          \/     \/               \/        \/      \_/
#
#

import re
from copy import copy

from jinja2 import Template
from openpyxl import load_workbook
from openpyxl.cell import Cell, MergedCell
from reporting import debug_keys, read_report
from utils import cell_value, create_sheet


def process_excel_template(filename, metadata=None, debug=False):
    if not metadata:
        metadata = {}

    report_data = read_report(filename)
    metadata.update(report_data)

    wb = load_workbook(filename=filename, data_only=True)
    for _, sheet in enumerate(wb):

        start = None
        end = None
        template = []
        style = []

        for i, row in enumerate(sheet):
            row_values = [cell_value(c) for _, c in enumerate(row) if isinstance(c, (Cell, MergedCell))]
            raw_values = " ".join(row_values)
            if "{%tr " in raw_values:
                logic_statement = "".join(row_values).replace("{%tr", "{%").replace(" %}", " -%}")
                if "{%tr for" in raw_values:
                    start = i
                elif "{%tr endfor" in raw_values:
                    end = i

                template.append(logic_statement)
            elif start is not None and end is None:
                style = [
                    {
                        "font": copy(c.font),
                        "border": copy(c.border),
                        "fill": copy(c.fill),
                        "number_format": copy(c.number_format),
                        "protection": copy(c.protection),
                        "alignment": copy(c.alignment),
                    }
                    for c in row
                ]
                template.append("-=+".join(row_values))

        if not template or start is None or end is None:
            continue

        raw_template = "\n".join(template)
        if debug:
            print("Raw Template:\n" + raw_template.replace("-=+", ""))

        jinga_template = Template(raw_template)
        processed = jinga_template.render(metadata)

        sheet.delete_rows(start + 1, amount=end - start + 1)

        for i, row in enumerate(processed.splitlines()):
            for j, c in enumerate(row.split("-=+")):
                cell = sheet.cell(row=start + i + 1, column=j + 1)
                cell.value = re.sub(r"\n\n+", "\n\n", c.replace("<w:br/>", "\n")).strip()
                try:
                    cell.font = copy(style[j]["font"])
                    cell.border = copy(style[j]["border"])
                    cell.fill = copy(style[j]["fill"])
                    cell.number_format = copy(style[j]["number_format"])
                    cell.protection = copy(style[j]["protection"])
                    cell.alignment = copy(style[j]["alignment"])
                except IndexError:
                    pass

    if debug and metadata:
        debugging_keys = debug_keys(metadata)
        debugging_keys.sort()

        create_sheet(wb, "Debug Table")

        debug_sheet = wb["Debug Table"]
        for i, key in enumerate(debugging_keys):
            cell = debug_sheet.cell(row=i + 1, column=1)
            cell.value = key

    wb.save(filename)
